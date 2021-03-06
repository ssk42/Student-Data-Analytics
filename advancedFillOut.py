###Searchable student analytics db built by Steve Reitz
##Python 3.6 using openpyxl module
##
##
##
##
## TODO:
# 4. Simple searches = saved?
## 1. Add searching by state
# 2. Daily Progress Report
# 5. Make it searchable by name too
# 6. Add multiple searching capabilities
# 3.. Gradebook functionality
# 7. CSV reading functionality

import os, openpyxl, pprint, logging, ftplib, sys, traceback
from openpyxl import Workbook
logger=logging.getLogger('ftpuploader')
agencyName=""
##Master sheet
#wbTest= openpyxl.load_workbook('test.xlsx')
#sheet= wbTest.get_sheet_by_name('Complete')
##Evaluation sheet
wbCompletion= openpyxl.load_workbook('test2.xlsx')
sheet2= wbCompletion.get_sheet_by_name('Complete')
##NLI Sheet.
##NOTE: You will always have to save it as a xlsv as when you download it from BQ, it's a csv.
wbNLI=openpyxl.load_workbook('test3.xlsx')
sheet3= wbNLI.get_sheet_by_name('Students-Never-Logged-into-Blue')
##Master Sheet with all of the states
wbStates=openpyxl.load_workbook('test5.xlsx')
sheet= wbStates.get_sheet_by_name('Master')
wbNew= Workbook()
ws=wbNew.active
wbAgency= Workbook()
wsAgency=wbAgency.active
##Dictionaries that store everything
student={}
userNLI={}
emailCheck={}
agency={}
completeCount=0
loggedInCount=0

completedUser=""
actualAgencyName=""
<<<<<<< HEAD

=======
q=0
qMax= wbStates['Querétaro'].max_row-1
j=0
jMax= wbStates['Jalisco'].max_row-1
c=0
cMax= wbStates['Ciudad de México'].max_row-1
a=0
aMax= wbStates['Aguascalientes'].max_row-1
t=0
tMax= wbStates['Tabasco'].max_row-1
co=0
coMax= wbStates['Colima'].max_row-1
ba=0
baMax= wbStates['Baja California'].max_row-1
si=0
siMax= wbStates['Sinaloa'].max_row-1
mo=0
moMax= wbStates['Morelos'].max_row-1
nl=0
nlMax= wbStates['Nuevo León'].max_row-1
slp=0
slpMax= wbStates['San Luis Potosí'].max_row-1
g=0
gMax= wbStates['Guanajuato'].max_row-1
o=0
oMax= wbStates['Oaxaca'].max_row-1
theRestCount=0
theRestCountTotal=407
>>>>>>> 4ab7ca9479d9209b740b097828c461e7ed60aa81

#test
#test2
#test3
##This next loop fills the dictionaries
for row in range(1, 427):
        user= sheet3['D'+str(row)].value
        userNLI.setdefault(user,{})
        email= sheet2['C'+str(row)].value
        evalEmail=sheet2['F'+str(row)].value
        emailCheck.update({email: evalEmail})

##This next loop fills out the sheet ala Google sheets but without any weird sorting, b/c yay Python
sheet=0
for sheet in wbStates:
        loggedInCount=0
        completeCount=0
        for rowNum in range(1,427):
            allFirstNames= sheet['A'+str(rowNum)].value
            allLastNames= sheet['B'+str(rowNum)].value
            if(sheet['C'+str(rowNum)].value!=None):
                    allAgencies= sheet['C'+str(rowNum)].value
                    agency.setdefault(allAgencies,{})
            allEmails= sheet['D'+str(rowNum)].value
            allUser= sheet['E'+str(rowNum)].value
            #Sets up the searchable categories
            student.setdefault(allUser,{"First Name": allFirstNames,"Last Name": allLastNames,"Agency": allAgencies, "User": allUser, "Email": allEmails, "¿Ha accedido al curso?": '',"¿Completó el curso?":''})
            yesOrNo= sheet.cell(row=rowNum, column=7)
            if allUser in userNLI:
                yesOrNo.value= 'No'
                student[allUser].update({"¿Ha accedido al curso?":'No'})
            else:
                yesOrNo.value= 'Sí'
                student[allUser].update({"¿Ha accedido al curso?":'Sí'})
                loggedInCount=loggedInCount+1
            if allEmails in emailCheck:
                sheet.cell(row=rowNum, column=7).value= 'Completó'
                student[allUser].update({"¿Completó el curso?":'Completó'})
                completeCount=completeCount+1
                #print(completeCount)
            elif( yesOrNo.value == 'Sí'):
                sheet.cell(row=rowNum, column=7).value= 'En progreso'
                student[allUser].update({"¿Completó el curso?":'En progreso'})
            else:
                sheet.cell(row=rowNum, column=7).value= 'No ha iniciado sesión'
                student[allUser].update({"¿Completó el curso?":'No ha iniciado sesión'})

for students in student:
<<<<<<< HEAD
        agency.update({student[students]['Agency']:student[students]['User']})
                      

#Searches for
=======
        agency.update({student[students]['Agency']:students})

#Searches for 
>>>>>>> 4ab7ca9479d9209b740b097828c461e7ed60aa81
def saveUserSearch(saveName, dictTerm, dictTerm2):
        print("Would you like to save this list?")
        yesOrNo=input()
        if yesOrNo=="Yes":
                wbNew= Workbook()
                ws= wbNew.active
                ws.title= str(saveName)
                for k,v in student[dictTerm2].items():
                        for row in range(1,(len(student[dictTerm2])+1)):
                                row2=row-1
                                ws['A'+str(row)].value= list(student[dictTerm2].keys())[row2]
                                ws['B'+str(row)].value= list(student[dictTerm2].values())[row2]
        wbNew.save(str(saveName)+'.xlsx')
        print("Saving file "+str(saveName)+'.xlsx')
        #end()

 #Searches for unsent evaluations
def saveUnsentEvals(saveName, dictTerm):
        print("Would you like to save this list?")
        yesOrNo=input()
        wbNew= Workbook()
        ws= wbNew.active
        ws.title= str(saveName)
        if yesOrNo=="Yes":
                for k,v in emailCheck.items():
                        for row in range(1,(len(emailCheck)+1)):
                                if values==None:
                                        row2=row-1
                                        ws['A'+str(row)].value= list(emailCheck)[row2]
                                        ws['B'+str(row)].value= list(emailCheck)[row2]
                wbNew.save(str(saveName)+'.xlsx')
                print("Saving file "+str(saveName)+'.xlsx')
        else:
                print("")
        end()

def agencyWrite(actualAgencyName, numRow, k, a):
        if a==1:
                firstName=student[k]['First Name']
                lastName=student[k]['Last Name']
                email=student[k]['Email']
                wsAgency.title=actualAgencyName
                if actualAgencyName in student[k]['Agency']:
                        ws['A'+str(numRow)].value= firstName
                        ws['B'+str(numRow)].value= lastName
                        ws['C'+str(numRow)].value= email
                        if wsAgency['D'+str(numRow)].value!=wsAgency.title:
                                wsAgency['A'+str(numRow)].value= None
                                wsAgency['B'+str(numRow)].value= None
                                wsAgency['C'+str(numRow)].value= None
                                wsAgency['D'+str(numRow)].value= None
                                wsAgency['E'+str(numRow)].value= None
                        ws['E'+str(numRow)].value= numRow
                        #print(firstName+''+student[k]['Agency'])
        else:
                firstName=student[k]['First Name']
                lastName=student[k]['Last Name']
                email=student[k]['Email']
                wsAgency.title=actualAgencyName
                if wsAgency.title==actualAgencyName:
                        wsAgency['A'+str(numRow)].value= firstName
                        wsAgency['B'+str(numRow)].value= lastName
                        wsAgency['C'+str(numRow)].value= email
                        wsAgency['D'+str(numRow)].value= actualAgencyName
                        if str(wsAgency['D'+str(numRow)].value)!=str(wsAgency.title):
                                wsAgency['A'+str(numRow)].value= None
                                wsAgency['B'+str(numRow)].value= None
                                wsAgency['C'+str(numRow)].value= None
                                wsAgency['D'+str(numRow)].value= None
                                wsAgency['E'+str(numRow)].value= None
                        else:
                                wsAgency['E'+str(numRow)].value= numRow

##These next three func
def agencyCompletionResults(agencyName):
<<<<<<< HEAD
        ##DESC:
=======
        ##DESC: 
>>>>>>> 4ab7ca9479d9209b740b097828c461e7ed60aa81
        ##TODO: Completion,
        ##           Add in emails
        ##           Save it in an excel sheet
        ##           Remove Nonetype error?
        ##           Add in different completion levels?
        ##           Apply to agencyLoginResults and agencyEvalResults
        ##TODO:login, eval filters
        #print(agencyName)

        firstName=""
        lastName=""
<<<<<<< HEAD
        numRow=1
=======
        numRow=1    
>>>>>>> 4ab7ca9479d9209b740b097828c461e7ed60aa81
        if agencyName=="All":
                #agencyName=""
                #initialAgencyName=agencyName
                for sheet in wbStates:
                        keyCounter=0
                        agencyName=sheet.title
                        for k,v in student.items():
                                if keyCounter<461:
                                        for k1,v1 in v.items():
                                                if v1=="Completó":
                                                        if k != None:
                                                                completedUser=k
                                                        #actualAgencyName=""
                                                        if agencyName=="Querétaro":
                                                                actualAgencyName=  'Jalpan'
                                                                if actualAgencyName in student[k]['Agency']:
<<<<<<< HEAD
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:
                                                                                agencyWrite(actualAgencyName, numRow, k,1)
                                                                                numRow=numRow+1
=======
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
                                                                                agencyWrite(actualAgencyName, numRow, k,1)
                                                                                numRow=numRow+1 
>>>>>>> 4ab7ca9479d9209b740b097828c461e7ed60aa81
                                                                        else:
                                                                                continue
                                                                actualAgencyName= agencyName
                                                                if actualAgencyName in student[k]['Agency']:
<<<<<<< HEAD
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:
=======
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
>>>>>>> 4ab7ca9479d9209b740b097828c461e7ed60aa81
                                                                                agencyWrite(actualAgencyName, numRow, k,1)
                                                                                #agencyCompletionResults(agencyName)
                                                                                numRow=numRow+1
                                                                        else:
                                                                                continue
<<<<<<< HEAD

                                                        elif agencyName=="Ciudad de México":
                                                                actualAgencyName= 'Distrito'
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:
=======
                                                      
                                                        elif agencyName=="Ciudad de México":
                                                                actualAgencyName= 'Distrito'
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
>>>>>>> 4ab7ca9479d9209b740b097828c461e7ed60aa81
                                                                                agencyWrite(actualAgencyName, numRow, k,1)
                                                                                #numRow=numRow+1
                                                                        else:
                                                                                continue

                                                                actualAgencyName= 'Ciudad de'
                                                                if actualAgencyName in student[k]['Agency']:
<<<<<<< HEAD
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:
=======
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
>>>>>>> 4ab7ca9479d9209b740b097828c461e7ed60aa81
                                                                                agencyWrite(actualAgencyName, numRow, k,1)
                                                                                #agencyCompletionResults(agencyName)
                                                                                numRow=numRow+1
                                                                        else:
                                                                                continue
                                                        else:
<<<<<<< HEAD
                                                                actualAgencyName=agencyName
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:
                                                                                agencyWrite(actualAgencyName, numRow, k,1)
                                                                                #agencyCompletionResults(actualAgencyName)
                                                                                numRow=numRow+1
                                                                        else:
                                                                                continue

                                keyCounter=keyCounter+1
                for sheet in wbStates:
                        keyCounter=0
                        agencyName=sheet.title
                        for k,v in student.items():
                                if keyCounter<463:
                                        for k1,v1 in v.items():
                                                if v1=="Completó":
                                                        if k != None:
                                                                completedUser=k
                                                        #actualAgencyName=""
                                                        if agencyName=="Querétaro":
                                                                actualAgencyName= agencyName
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:
                                                                                #agencyWrite(actualAgencyName, numRow, k,1)
                                                                                agencyCompletionResults(agencyName)

                                                                                #numRow=numRow+1
                                                                        else:
                                                                                continue

                                                        elif agencyName=="Ciudad de Mexico":
                                                                actualAgencyName= 'Ciudad de México'
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:
                                                                                #agencyWrite(actualAgencyName, numRow, k,1)
                                                                                agencyCompletionResults(agencyName)
                                                                                #numRow=numRow+1
                                                                        else:
                                                                                continue
                                                        else:
                                                                actualAgencyName=str(agencyName)
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:
                                                                                #agencyWrite(actualAgencyName, numRow, k,1)
                                                                                #print(agencyName)
                                                                                agencyCompletionResults(str(agencyName))
                                                                                continue

                                                                                #numRow=numRow+1
=======
                                                                actualAgencyName=agencyName                                                         
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
                                                                                agencyWrite(actualAgencyName, numRow, k,1)
                                                                                #agencyCompletionResults(actualAgencyName)
                                                                                numRow=numRow+1
>>>>>>> 4ab7ca9479d9209b740b097828c461e7ed60aa81
                                                                        else:
                                                                                continue

                                keyCounter=keyCounter+1
                for sheet in wbStates:
                        keyCounter=0
                        agencyName=sheet.title                        
                        for k,v in student.items():
                                if keyCounter<463:
                                        for k1,v1 in v.items():
                                                if v1=="Completó":
                                                        if k != None:
                                                                completedUser=k
                                                        #actualAgencyName=""
                                                        if agencyName=="Querétaro":
                                                                actualAgencyName= agencyName
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
                                                                                #agencyWrite(actualAgencyName, numRow, k,1)
                                                                                agencyCompletionResults(agencyName)
                                                                               
                                                                                #numRow=numRow+1
                                                                        else:
                                                                                continue
                                                      
                                                        elif agencyName=="Ciudad de Mexico":
                                                                actualAgencyName= 'Ciudad de México'
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
                                                                                #agencyWrite(actualAgencyName, numRow, k,1)
                                                                                agencyCompletionResults(agencyName)
                                                                                #numRow=numRow+1
                                                                        else:
                                                                                continue
                                                        else:
                                                                actualAgencyName=str(agencyName)                                                         
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
                                                                                #agencyWrite(actualAgencyName, numRow, k,1)
                                                                                #print(agencyName)
                                                                                agencyCompletionResults(str(agencyName))

                                                                                #numRow=numRow+1
                                                                        else:
                                                                                continue

                                #keyCounter=keyCounter+1                                
                wbNew.save('Completes sorted by Agency.xlsx')
                print('Completes sorted by Agency.xlsx')
        else:
                #initialAgencyName=agencyName
<<<<<<< HEAD

                numRow=1
                for k,v in student.items():
                        #if keyCounter<463:
=======
                keyCounter=0
                numRow=1
                #theRestCount=theRestCount
                for k,v in student.items():
                        #if keyCounter<463:                        
>>>>>>> 4ab7ca9479d9209b740b097828c461e7ed60aa81
                                for k1,v1 in v.items():
                                        if v1=="Completó":
                                                completedUser=k
                                                if agencyName=="Querétaro":
<<<<<<< HEAD
                                                        actualAgencyName= "Querétaro"
                                                        if actualAgencyName in student[k]['Agency']:
                                                                if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:
                                                                        agencyWrite(actualAgencyName, numRow, k,0)
                                                                        numRow=numRow+1
                                                                else:
                                                                        continue
                                                        actualAgencyName=  'Jalpan'
                                                        if actualAgencyName in student[k]['Agency']:
                                                                if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:
                                                                        agencyWrite(actualAgencyName, numRow, k,0)
                                                                        numRow=numRow+1
                                                                else:
                                                                        continue
                                                elif agencyName=="Ciudad de México":
                                                        actualAgencyName= 'Distrito'
                                                        if actualAgencyName in student[k]['Agency']:
                                                                if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:
                                                                        agencyWrite(actualAgencyName, numRow, k,0)
                                                                else:
                                                                        continue

                                                        actualAgencyName= 'Ciudad de México'
                                                        #print(actualAgencyName)
                                                        if actualAgencyName in student[k]['Agency']:
                                                                if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:
                                                                        agencyWrite(actualAgencyName, numRow, k,0)
                                                                        numRow=numRow+1
                                                                else:
                                                                        continue
                                                
                                                else:
                                                        actualAgencyName=str(agencyName)
                                                        if actualAgencyName in student[k]['Agency']:
                                                                if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:
                                                                        agencyWrite(actualAgencyName, numRow, k,0)
                                                                        numRow=numRow+1
                                                                        continue
                                                                else:
                                                                        continue

                        #keyCounter+=1
                wbAgency.save(agencyName+' Completes.xlsx')
                print("Saved "+agencyName+' Completes.xlsx')

=======
                                                        #print(wbStates['Querétaro'].max_row)
                                                        global q
                                                        if q<qMax:                                                                
                                                                actualAgencyName= "Querétaro"
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
                                                                                agencyWrite(actualAgencyName, numRow, k,0)
                                                                                numRow=numRow+1
                                                                                q=q+1
                                                                        else:
                                                                                continue
                                                                actualAgencyName=  'Jalpan'
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
                                                                                agencyWrite(actualAgencyName, numRow, k,0)
                                                                                numRow=numRow+1
                                                                                q=q+1
                                                                                print(str(q)+" q")
                                                                        else:
                                                                                continue  
                                                elif agencyName=="Ciudad de México":
                                                        #print(wbStates[agencyName].max_row)
                                                        global c
                                                        if c<cMax:                                                            
                                                                actualAgencyName= 'Distrito'
                                                                #print(actualAgencyName)
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
                                                                                agencyWrite(actualAgencyName, numRow, k,0)
                                                                                c=c+1
                                                                                #numRow=numRow+1
                                                                        else:
                                                                                continue

                                                                actualAgencyName= 'Ciudad de México'
                                                                #print(actualAgencyName)
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
                                                                                agencyWrite(actualAgencyName, numRow, k,0)
                                                                                numRow=numRow+1
                                                                                c=c+1
                                                                                print(str(c)+" c")
                                                                        else:
                                                                                continue
                                                elif agencyName=="Jalisco":
                                                        #print(wbStates[agencyName].max_row)
                                                        global j
                                                        if j<jMax:                                                            
                                                                actualAgencyName=str(agencyName)
                                                                #print(actualAgencyName)
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
                                                                                agencyWrite(actualAgencyName, numRow, k,0)
                                                                                numRow=numRow+1
                                                                                j=j+1
                                                                                print(str(j)+" j")
                                                                        else:
                                                                                continue
                                                elif agencyName=="Aguascalientes":
                                                        #print(wbStates[agencyName].max_row)
                                                        global a
                                                        if a<aMax:                                                            
                                                                actualAgencyName=str(agencyName)
                                                                #print(actualAgencyName)
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
                                                                                agencyWrite(actualAgencyName, numRow, k,0)
                                                                                numRow=numRow+1
                                                                                a=a+1
                                                                                print(str(a)+" a")
                                                                        else:
                                                                                continue
                                                elif agencyName=="Tabasco":
                                                        #print(wbStates[agencyName].max_row)
                                                        global t
                                                        if t<tMax:                                                            
                                                                actualAgencyName=str(agencyName)
                                                                #print(actualAgencyName)
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
                                                                                agencyWrite(actualAgencyName, numRow, k,0)
                                                                                numRow=numRow+1
                                                                                t=t+1
                                                                                print(str(t)+" t")
                                                                        else:
                                                                                continue
                                                elif agencyName=="Colima": #30
                                                        #print(wbStates[agencyName].max_row)
                                                        global co
                                                        if co<coMax:                                                            
                                                                actualAgencyName=str(agencyName)
                                                                #print(actualAgencyName)
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
                                                                                agencyWrite(actualAgencyName, numRow, k,0)
                                                                                numRow=numRow+1
                                                                                co=co+1
                                                                                print(str(co)+" co")
                                                                        else:
                                                                                continue
                                                elif agencyName=="Baja California": #34
                                                        #print(wbStates[agencyName].max_row)
                                                        global ba
                                                        if ba<baMax:                                                            
                                                                actualAgencyName=str(agencyName)
                                                                #print(actualAgencyName)
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
                                                                                agencyWrite(actualAgencyName, numRow, k,0)
                                                                                numRow=numRow+1
                                                                                ba=ba+1
                                                                                print(str(ba)+" ba")
                                                                        else:
                                                                                continue 
                                                elif agencyName=="Sinaloa": #18
                                                        #print(wbStates[agencyName].max_row)
                                                        global si
                                                        if si<siMax:                                                            
                                                                actualAgencyName=str(agencyName)
                                                                #print(actualAgencyName)
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
                                                                                agencyWrite(actualAgencyName, numRow, k,0)
                                                                                numRow=numRow+1
                                                                                si=si+1
                                                                                print(str(si)+" si")
                                                                        else:
                                                                                continue
                                                elif agencyName=="Morelos": #39
                                                        #print(wbStates[agencyName].max_row)
                                                        global mo
                                                        if mo<moMax:                                                            
                                                                actualAgencyName=str(agencyName)
                                                                #print(actualAgencyName)
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
                                                                                agencyWrite(actualAgencyName, numRow, k,0)
                                                                                numRow=numRow+1
                                                                                mo=mo+1
                                                                                print(str(mo)+" mo")
                                                                        else:
                                                                                continue
                                                elif agencyName=="Nuevo León": #38
                                                        #print(wbStates[agencyName].max_row)
                                                        global nl
                                                        if nl<nlMax:                                                            
                                                                actualAgencyName=str(agencyName)
                                                                #print(actualAgencyName)
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
                                                                                agencyWrite(actualAgencyName, numRow, k,0)
                                                                                numRow=numRow+1
                                                                                nl=nl+1
                                                                                print(str(nl)+" nl")
                                                                        else:
                                                                                continue
                                                elif agencyName=="San Luis Potosí": #47
                                                        #print(wbStates[agencyName].max_row)
                                                        global slp
                                                        if slp<slpMax:                                                            
                                                                actualAgencyName=str(agencyName)
                                                                #print(actualAgencyName)
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
                                                                                agencyWrite(actualAgencyName, numRow, k,0)
                                                                                numRow=numRow+1
                                                                                slp=slp+1
                                                                                print(str(slp)+" slp")
                                                                        else:
                                                                                continue
                                                elif agencyName=="Guanajuato": #45
                                                        #print(wbStates[agencyName].max_row)
                                                        global g
                                                        if g<gMax:                                                            
                                                                actualAgencyName=str(agencyName)
                                                                #print(actualAgencyName)
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
                                                                                agencyWrite(actualAgencyName, numRow, k,0)
                                                                                numRow=numRow+1
                                                                                g=g+1
                                                                                print(str(g)+" g")
                                                                        else:
                                                                                continue
                                                elif agencyName=="Oaxaca": #44
                                                        #print(wbStates[agencyName].max_row)
                                                        global o
                                                        if o<oMax:                                                            
                                                                actualAgencyName=str(agencyName)
                                                                #print(actualAgencyName)
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
                                                                                agencyWrite(actualAgencyName, numRow, k,0)
                                                                                numRow=numRow+1
                                                                                o=o+1
                                                                                print(str(o)+" o")
                                                                        else:
                                                                                continue                                                                         
                                                else:
                                                        #print(theRestCountTotal)
                                                        global theRestCount
                                                        if theRestCount<10:
                                                                actualAgencyName=str(agencyName)
                                                                #print(actualAgencyName)
                                                                if actualAgencyName in student[k]['Agency']:
                                                                        if student[k]['First Name'] and student[k]['Last Name'] and student[k]['Email'] != None:                                                                
                                                                                agencyWrite(actualAgencyName, numRow, k,0)
                                                                                numRow=numRow+1
                                                                                theRestCount=theRestCount+1
                                                                                print(str(theRestCount)+ " theRestCount "+ actualAgencyName)                                                                       
                                                                        else:
                                                                                continue
                wbAgency.save(agencyName+' Completes.xlsx')
                print("Saved "+agencyName+' Completes.xlsx')
                
>>>>>>> 4ab7ca9479d9209b740b097828c461e7ed60aa81
def agencyLoginResults(agencyName):
        print(agencyName)
def agencyEvalResults(agencyName):
        print(agencyName)

def generalSearch():
##DESC: This next part is the bit that allows you to search for users. Aka the functionality that any LMS should have
        print("What would you like to know? Would you like to search a user? Would you like to see who hasn't received an evaluation?")
        searchSpec=input()
        if 'user' in searchSpec:
                print('What username would you like to search for?')
                userSearch=input()
                if userSearch in student.keys():
                    print("What would you like to know? \n You can answer with \n Everything, Usename, Eval sent?, Email, Logged In, Complete, or Agency")
                    specify=input()
                    if(specify=='Everything'):
                        pprint.pprint(student[userSearch])
                        saveUserSearch(str(userSearch)+ " User Info", userSearch, str(userSearch))
                    elif(specify=='Username'):
                        print(student[userSearch]['User'])
                    elif(specify=='Email'):
                        print(student[userSearch]['Email'])
                        #saveSpecificUserSearch(str(userSearch)+ " Username", userSearch, str(userSearch),'Email')
                    elif(specify=='Logged In'):
                        print(student[userSearch]['¿Ha accedido al curso?'])
                        #saveSpecificUserSearch(str(userSearch)+ " Username", userSearch, str(userSearch),'¿Ha accedido al curso?')
                    elif(specify=='Complete'):
                        print(student[userSearch]['¿Completó el curso?'])
                    elif(specify=='Agency'):
                        print(student[userSearch]['Agency'])
                    elif(specify=='Eval sent?'):
                        if(emailCheck[(student[userSearch]['Email'])] is None):
                           print('No')
                        else:
                           print(emailCheck[(student[userSearch]['Email'])])
                else:
                    print("Sorry, user not found.")
        elif 'eval' in searchSpec:
                wbNew= Workbook()
                ws= wbNew.active
                for k,v in emailCheck.items():
                        if v==None:
                            print(k)
                numRow=1
                for row in range(1,(len(emailCheck)+1)):
                        row2=(row-1)
                        if list(emailCheck.values())[row2]==None:
                                ws['A'+str(numRow)].value= list(emailCheck.keys())[row2]
                                ws['B'+str(numRow)].value= list(emailCheck.values())[row2]
                                numRow=numRow+1
                        else:
                                 row=row+1
                wbNew.save(' No Evals sent.xlsx')
                print("This list can also be found as  "+'No Evals sent.xlsx in ' +os.getcwd())
                                    #saveUnsentEvals(" Evals Not Sent Yet", None)
        elif 'agency' in searchSpec:
                print('What agency would you like to search for?')
                agencySearch=input()
                print('Would you like to filter search results for this agency?\n By completion? By logged in? By evaluation sent?')
                filterSearch=input()
                if(filterSearch=="Completion"):
                        agencyCompletionResults(agencySearch)
                elif(filterSearch=="Logged in"):
                        agencyLoginResults(agencySearch)
                elif("eval" in filterSearch):
                        agencyEvalResults(agencySearch)
        #end()

print('Updated all known student info data as updated.xlsx')
wbStates.save('updated.xlsx')
print('Would you like to make another search?')
repeatAnswer=input()
if repeatAnswer=='Yes':
        generalSearch()
if repeatAnswer=='No':
        quit


generalSearch()
##Saves it as a Excel sheet
##print('Saved updated.xlsx')
##wbStates.save('updated.xlsx')
##print('Would you like to make another search?')
##repeatAnswer=input()
##if repeatAnswer=='Yes':
##        generalSearch()
##if repeatAnswer=='No':
##        quit
